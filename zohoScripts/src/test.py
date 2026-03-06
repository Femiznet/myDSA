import sys
import json
import requests
import logging as log
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from requests.exceptions import HTTPError, Timeout, ConnectionError
from utils import required
from watch import run_watcher, file_queue
import time
import argparse

# 1. Setup CLI Arguments
parser = argparse.ArgumentParser(description="Zoho Ticket Automation")
parser.add_argument("--refresh", action="store_true", help="Bypass cache and force API pull")
parser.add_argument("--dry", action="store_true", help="Generate Excel files but do NOT trigger emails")
args = parser.parse_args()

# 2. Setup Paths & Start Watcher
cache_dir = Path(r"C:\Users\DELL\devFiles\zohoScripts\cache")
cache_dir.mkdir(exist_ok=True)
if not args.dry: run_watcher()

log.basicConfig(
    filename=r'..\log\get_ticket.log',
    level=log.INFO,
    format='%(levelname)s - %(message)s - %(asctime)s',
    filemode='a'
)

# Configuration
URL = "https://desk.zoho.com/api/v1/tickets/search"
HEADER = {
    "Authorization": f"Zoho-oauthtoken {required('AUTH-TOKEN')}",
    "orgId": required("ORG-ID")
}

# Settings
from_date, end_date = "1 may 2025", "4 mar 2026"
mayowaData, olawunmiData = [], []
use_cache = False
cache_files = {"Estate": cache_dir / "Estate.json", "Plant": cache_dir / "Plant.json"}

# --- 3. Robust Cache Logic ---
if not args.refresh:
    if all(f.exists() and f.stat().st_size > 5 for f in cache_files.values()):
        mtime = min(f.stat().st_mtime for f in cache_files.values())
        time_diff = (time.time() - mtime) / 60 

        if time_diff < 10:
            try:
                with open(cache_files["Estate"], "r") as f: mayowaData = json.load(f)
                with open(cache_files["Plant"], "r") as f: olawunmiData = json.load(f)
                if mayowaData or olawunmiData:
                    print(f"Cache is fresh ({int(time_diff)} mins old). Loading data...")
                    use_cache = True
            except Exception:
                log.error("Cache file corrupted. Falling back to API.")

# --- 4. Data Collection (API) ---
if not use_cache:
    print("Calling Zoho API for fresh data...")
    status, offset, limit = "open", 0, 100

    while True:
        PARAM = {"from": offset, "limit": limit, "status": status}
        try:
            response = requests.get(url=URL, headers=HEADER, params=PARAM, timeout=30)
            
            if 'application/json' not in response.headers.get('Content-Type', ''):
                log.error(f"API Error: Server sent non-JSON response (Status {response.status_code})")
                print(f"Response Snippet: {response.text[:200]}")
                break

            response.raise_for_status()
            jsonData = response.json()
        except Exception as err:
            log.error(f"API Connection Error: {err}")
            break
        
        tickets = jsonData.get("data", [])
        if not tickets: break

        for data in tickets:
            try:
                assignee = data.get("assignee") or {"firstName": "Mayowa", "lastName": "Ajuwon"}
                fName, lName = assignee.get("firstName", ""), assignee.get("lastName", "")
                
                ptr = None
                if fName.lower() == "mayowa" and lName.lower() == "ajuwon": ptr = mayowaData
                elif fName.lower() == "olawunmi" and lName.lower() == "abiodun": ptr = olawunmiData
                else: continue

                desc = BeautifulSoup(data.get("description", "") or "", "html.parser").get_text() 
                fields = data.get("customFields", {})
                cov = fields.get("Request Coverage", "")
                
                if ptr is olawunmiData and cov.lower() != "plant maintenance": continue
                
                ptr.append({
                    "TicketOwner": f"{fName} {lName}", "TicketId": data.get("ticketNumber"),
                    "Subject": data.get("subject"), "Description": desc,
                    "CreatedTime": data.get("createdTime"), "ClosedTime": data.get("closedTime"),
                    "Status": data.get("statusType"), "Address": fields.get("Address"),
                    "Location": fields.get("Location"), "rCoverage": cov,
                    "rCategory": fields.get("Request Category"),
                })
            except: continue
            
        offset += limit
        print(f"Collecting Data (Offset: {offset}).....")

    if mayowaData or olawunmiData:
        for loc, data_list in [("Estate", mayowaData), ("Plant", olawunmiData)]:
            with open(cache_files[loc], "w") as f:
                json.dump(data_list, f, indent=3)

# --- 5. Excel Generation ---
allData = [mayowaData, olawunmiData]
locations = ["Estate", "Plant"]

for i, uniqueData in enumerate(allData):
    if not uniqueData:
        print(f"No data for {locations[i]}, skipping Excel.")
        continue

    df = pd.DataFrame(uniqueData)
    df["CreatedTime"] = pd.to_datetime(df["CreatedTime"], utc=True)
    start, end = pd.to_datetime(from_date, utc=True), pd.to_datetime(end_date, utc=True)
    
    # Filter based on dates
    df = df[df["CreatedTime"].between(start, end)].copy()
    df["CreatedTime"] = df["CreatedTime"].dt.strftime("%d %b %Y")

    # Path Setup
    today = datetime.now().strftime("%Y-%m-%d")
    fpath = Path(fr"C:\Users\DELL\devFiles\tickets\{locations[i]}Tickets")
    fpath.mkdir(parents=True, exist_ok=True)
    
    filePath = fpath / f"{locations[i]}_{today}_ticket.xlsx"
    df.to_excel(filePath, index=False)
    
    # Styling Excel
    wb = load_workbook(filePath)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
        # Using .column_letter attribute of the first cell in the tuple
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    wb.save(filePath)
    print(f"Excel generated: {filePath.name}")

# --- 6. Finalise ---
log.info("All operations completed successfully")

if not args.dry:
    print("Giving watcher a moment to detect files...")
    time.sleep(2) 

    timeout, start_wait = 90, time.time() # Increased timeout to 90s for safety
    print(f"Waiting for emails to send (In Queue: {file_queue.unfinished_tasks})...")

    while file_queue.unfinished_tasks > 0:
        if time.time() - start_wait > timeout:
            print("Timeout reached! Closing script before all emails sent.")
            break
        time.sleep(1) 
else:
    print("Dry run complete. No emails were queued.")

print("Process finished. Exiting.")
