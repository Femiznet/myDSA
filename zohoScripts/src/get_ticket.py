import sys
import json
import requests
import logging as log
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from requests.exceptions import HTTPError, Timeout, ConnectionError
from utils import required
from watch import run_watcher, file_queue
import time

# start watching folder for new files
run_watcher()

log.basicConfig(
    filename=r'..\log\get_ticket.log',
    level=log.INFO,
    format='%(levelname)s - %(message)s - %(asctime)s',
    filemode='a'
)

# Configuration
URL = "https://desk.zoho.com/api/v1/tickets/search"
TIME = 30
HEADER = {
    "Authorization": f"Zoho-oauthtoken {required('AUTH-TOKEN')}",
    "orgId": required("ORG-ID")
    }

def param(offset=0, limit=1, **kwargs):
    parameters = {"from": offset, "limit": limit} | kwargs
    return parameters

# Input Settings
status = "open"
from_date = "1 may 2025"
end_date = "4 mar 2026"

mayowaData = []
olawunmiData = []
offset = 0
limit = 100
error = None

# 1. Data Collection Loop
while True:
    PARAM = param(offset=offset, limit=limit, status=status)
    
    try:
        response = requests.get(url=URL, headers=HEADER, params=PARAM, timeout=TIME)
        response.raise_for_status()
        jsonData = response.json()
    except (ConnectionError, Timeout) as err:
        log.error(f"Connection timed out: {err}")
        sys.exit()
    except HTTPError as err:
        log.error(f"An error occurred: {err}")
        error = True
    except json.JSONDecodeError:
        break
    
    if error:
        log.error(f"Response Content: {response.content}")
        sys.exit()
    
    # Check if data exists in response
    tickets = jsonData.get("data", [])
    if not tickets:
        break

    for data in tickets:
        try:
            assignee = data.get("assignee")
            if not assignee:
                assignee = {"firstName": "Mayowa", "lastName": "Ajuwon"}
        except AttributeError:
            log.error(f"Ticket => {data.get('ticketNumber')} has No assignee")
            continue
        
        firstName = assignee.get("firstName", "")
        lastName = assignee.get("lastName", "")
        
        # Point to correct list
        Data_ptr = None
        if firstName.lower() == "mayowa" and lastName.lower() == "ajuwon":
            Data_ptr = mayowaData
        elif firstName.lower() == "olawunmi" and lastName.lower() == "abiodun":
            Data_ptr = olawunmiData
        else:
            continue

        # Parse Description
        text = data.get("description", "") or ""
        desc = BeautifulSoup(text, "html.parser").get_text() 
        
        customFields = data.get("customFields", {})
        coverage = customFields.get("Request Coverage", "")
        
        # Coverage Filter for Olawunmi
        if Data_ptr is olawunmiData:
            if coverage.lower() != "plant maintenance":
                continue
        
        full_data = {
            "TicketOwner": f"{firstName} {lastName}",
            "TicketId": data.get("ticketNumber"),
            "Subject": data.get("subject"),
            "Description": desc,
            "CreatedTime": data.get("createdTime"),
            "ClosedTime": data.get("closedTime"),
            "Status": data.get("statusType"),
            "Address": customFields.get("Address"),
            "Location": customFields.get("Location"),
            "rCoverage": coverage,
            "rCategory": customFields.get("Request Category"),
        }
        Data_ptr.append(full_data)
        
    offset += limit
    print(f"Collecting Data (Offset: {offset}).....")

# 2. Data Processing & Excel Generation
cache = Path(r"C:\Users\DELL\devFiles\zohoScripts\cache")
cache.mkdir(exist_ok=True)
allData = [mayowaData, olawunmiData]
location = ["Estate", "Plant"]

for i, uniqueData in enumerate(allData):
    if not uniqueData:
        print(f"No data for {location[i]}, skipping...")
        continue

    # Cache to JSON
    file = Path(cache, f"{location[i]}.json")
    with open(file, "w") as jsonFile:
        json.dump(uniqueData, jsonFile, indent=3)

    # DataFrame Operations
    df = pd.DataFrame(uniqueData)
    df["CreatedTime"] = pd.to_datetime(df["CreatedTime"], utc=True)
    
    start = pd.to_datetime(from_date, utc=True)
    end = pd.to_datetime(end_date, utc=True)
    
    # Filter and format
    df = df[df["CreatedTime"].between(start, end)].copy()
    df["CreatedTime"] = df["CreatedTime"].dt.strftime("%d %b %Y")

    # Path Setup
    today = datetime.now().date()
    fpath = Path(fr"C:\Users\DELL\devFiles\tickets\{location[i]}Tickets")
    fpath.mkdir(parents=True, exist_ok=True)
    
    filePath = fpath / f"{location[i]}{today}ticket.xlsx"
    df.to_excel(filePath, index=False)
    
    print(df)
    
    # 3. Excel Styling
    wb = load_workbook(filePath)
    ws = wb.active

    # Alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column Widths
    for col in ws.columns:
        max_len = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_len + 2, 50)

    wb.save(filePath)
    
log.info("All operations completed")

# Timeout settings (60 seconds)
timeout = 60 
start_wait = time.time()

print("Waiting for emails to send (Timeout: 60s)...")

# Proper Queue Check: check if queue is empty AND all tasks are finished
while file_queue.unfinished_tasks > 0:
    if time.time() - start_wait > timeout:
        print("Timeout reached! Closing script before all emails sent.")
        log.warning("Script timed out before queue was empty.")
        break
    time.sleep(1) 

print("Process finished. Exiting.")
